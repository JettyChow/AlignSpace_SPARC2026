"""
Seed the Postgres tables from `alignspace dataset.xlsx`.

The workbook is a single sheet with labeled sections (FIRMS, ROOM_TYPES,
BUDGETS, STYLES, IMAGES, IMAGES_STYLES, PRESETS, PRESET_STYLES, MATERIALS,
ITEMS, ITEMS_MATERIALS, PRESET_ITEMS). Each section has a marker row, a
header row of column names, then data rows until the next blank row.

This script parses each section and upserts (insert-or-update, by primary
key) rows into the matching SQLAlchemy model/table. Safe to re-run — rows
are merged by primary key, not duplicated.

Usage (from as-ai-server/):
    pip install -r requirements.txt
    pip install openpyxl
    export DATABASE_URL=postgresql+psycopg2://user:pass@host:5432/dbname
    python scripts/seed_from_xlsx.py "../alignspace dataset.xlsx"
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from sqlalchemy import Boolean, Integer

# Make "app" importable when run as a plain script (same trick as migrations/env.py).
sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "src"))

from app.database import SessionLocal  # noqa: E402
from app.models import (  # noqa: E402
    Budget,
    Firm,
    Image,
    ImageStyle,
    Item,
    ItemMaterial,
    Material,
    Preset,
    PresetItem,
    PresetStyle,
    RoomType,
    Style,
)

# Section marker (as it appears in column A) -> model class.
# Order matters: parents must be seeded before children that FK-reference them.
SECTIONS = [
    ("FIRMS", Firm),
    ("ROOM_TYPES", RoomType),
    ("BUDGETS", Budget),
    ("STYLES", Style),
    ("IMAGES", Image),
    ("IMAGES_STYLES", ImageStyle),
    ("PRESETS", Preset),
    ("PRESET_STYLES", PresetStyle),
    ("MATERIALS", Material),
    ("ITEMS", Item),
    ("ITEMS_MATERIALS", ItemMaterial),
    ("PRESET_ITEMS", PresetItem),
]

# xlsx header -> model column name, per section (only needed where they differ).
HEADER_RENAMES = {
    "ITEMS": {"Set": "item_set"},
}


def _coerce(model, row: dict, section: str) -> dict:
    """Cast values to the type the model column expects (xlsx gives floats/bools)."""
    renames = HEADER_RENAMES.get(section, {})
    row = {renames.get(k, k): v for k, v in row.items()}

    columns = {c.name: c for c in model.__table__.columns}
    out = {}
    for key, value in row.items():
        col = columns.get(key)
        if col is None:
            continue  # drop columns not present on the model (notes, helper cols, etc.)
        if value is not None:
            if isinstance(col.type, Integer):
                value = int(value)
            elif isinstance(col.type, Boolean):
                value = bool(value)
        out[key] = value
    return out


def parse_sections(path: str) -> dict:
    """Return {section_name: [row_dict, ...]} parsed from the workbook."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    markers = {name for name, _ in SECTIONS}
    sections: dict[str, list[dict]] = {}

    i = 0
    while i < len(rows):
        first_cell = rows[i][0] if rows[i] else None
        if first_cell in markers:
            section_name = first_cell
            header_idx = i + 1
            while header_idx < len(rows) and not any(v is not None for v in rows[header_idx]):
                header_idx += 1
            header = rows[header_idx]

            data_rows = []
            j = header_idx + 1
            while j < len(rows) and any(v is not None for v in rows[j]):
                row_dict = {
                    header[k]: rows[j][k]
                    for k in range(len(header))
                    if header[k] is not None
                }
                data_rows.append(row_dict)
                j += 1

            sections[section_name] = data_rows
            i = j
        else:
            i += 1

    return sections


def seed(path: str) -> None:
    sections = parse_sections(path)
    session = SessionLocal()
    try:
        for name, model in SECTIONS:
            rows = sections.get(name, [])
            for raw_row in rows:
                clean = _coerce(model, raw_row, name)
                session.merge(model(**clean))
            session.flush()
            print(f"{name}: upserted {len(rows)} row(s)")
        session.commit()
        print("Done.")
    except Exception:
        session.rollback()
        raise
    finally:
        session.close()


if __name__ == "__main__":
    xlsx_path = sys.argv[1] if len(sys.argv) > 1 else "../alignspace dataset.xlsx"
    seed(xlsx_path)
